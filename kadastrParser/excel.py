from openpyxl import load_workbook


class Excel:
    def __init__(self, filename):
        self.wb = load_workbook(filename=filename)
        self.filename = filename

    def get_col(self, col='A'):
        active_sheet = self.wb.active
        kad_numbers = list(active_sheet[col])
        for i in range(len(kad_numbers)):
            kad_numbers[i] = kad_numbers[i].value
        return kad_numbers

    def write(self, cords, value):
        active_sheet = self.wb.active
        active_sheet.cell(row=cords[0], column=cords[1]).value = value
        return True

    def close(self):
        self.wb.save(filename=self.filename)
        self.wb.close()


class MyCell:
    ALPHABET = ["A", "B", "C", "D", "E", "F", "G", "H",
                "I", "J", "K", "L", "M", "N", "O", "P",
                "Q", "R", "S", "T", "U", "V", "W", "X",
                "Y", "Z"]

    def __init__(self, file, sheet, c, r):
        self.file = file
        self.sheet = sheet
        self.c = c
        self.r = r

    def info(self):
        return self.file + " " + self.sheet + " " + self.refCol() + self.r

    def refCol(self):
        alphSize = len(self.ALPHABET)
        cname = int(self.c) % alphSize
        doubles = int(self.c) // alphSize + 1
        cname = (self.ALPHABET[cname - 1]) * doubles
        return cname
